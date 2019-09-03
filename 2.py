from bs4 import BeautifulSoup
from urllib.request import Request, urlopen
from selenium import webdriver


#load settings from json
import json
with open('set.json', 'r', encoding='utf-8') as fh:
    arr_set = json.load(fh)
with open('set_name.json', 'r', encoding='utf-8') as fh:
    arr_set_name = json.load(fh)
with open('set_param_oit.json', 'r', encoding='utf-8') as fh:
    arr_par_o = json.load(fh)
with open('set_update.json', 'r', encoding='utf-8') as fh:
    arr_update = json.load(fh)



#Get html from url
def gethtml(url):
    req = Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    driver = webdriver.Chrome(chrome_options=options)
    driver.get(url)
    ghtml = driver.page_source
    driver.quit()
    return ghtml

#get country
def getcountry(lnk, num):
    res = ''
    cntrlink = lnk.contents[num]
    imglink = cntrlink.findAll('img')
    if imglink:
        for fimg in imglink:
            imgnm = fimg.get('alt')
            if imgnm:
                if res != '':
                    res = res + ', '
                res = res + imgnm
    return res

#find position of characteristic in excel according to text of characteristic
def find_charact(charact_text):
    res = ''
    if charact_text:
        for pos, txt in arr_set_name.items():
            if txt == charact_text:
                for chr, arr in arr_set.items():
                    if int(pos) in arr:
                        res = chr
    return res

#find position of oit parametres in excel
def find_charact_par_o(charact_text):
    res = ''
    if charact_text:
        for pos, txt in arr_par_o.items():
            if txt == charact_text:
                res = pos
    return res

import re
#update value
def update_value_par(charact_text, value):
    res = value
    if value:
        pos = find_charact_par_o(charact_text)
        if pos != '':
            for pos2, mult_update in arr_update.items():
                if pos2 == pos:
                    all_floats = re.findall('\d+\.?\d*', value)
                    for find_float in all_floats:
                        new_float = float(find_float) * float(mult_update)
                        if new_float.is_integer():
                            new_float = int(new_float)
                        res = str(res).replace(str(find_float), str(new_float))
    return res







generated_html = gethtml("http://......site.html")

#Excel
from openpyxl import load_workbook
wb = load_workbook('temp.xlsx')
ws = wb.active

import string

rowbegin_1 = 3
currow = 1
bool_items = False

#parse html
soup = BeautifulSoup(generated_html, 'html.parser')

for link in soup.find_all('table'):
    for trlink in link.find_all('tr'):
        if trlink:
            tdlink = trlink.contents[1]
            alink = tdlink.find('a')
            if alink:
                currow += 1
                adzz = alink.get('title')
                hrefdzz = alink.get('href')
                if adzz:
                    res = adzz
                    print(res)
                    #f.write(res+'\n')
                    ws['B' + str(currow)] = res              	     #name
                    ws['C' + str(currow)] = trlink.contents[4].text  #date 
                    ws['F' + str(currow)] = trlink.contents[3].text  #own 
                    ws['G' + str(currow)] = trlink.contents[2].text  #full 
                    ws['O' + str(currow)] = getcountry(trlink, 8)    #country 
                    ws['R' + str(currow)] = getcountry(trlink, 9)    #get 



                    #go to the next inner level
                    secondpage = gethtml("http://.....site.org/" + hrefdzz)
                    second_soup = BeautifulSoup(secondpage, 'html.parser')

                    #get type of operator
                    fnd = second_soup.find(text="Тип")
                    if fnd:
                        ws['P' + str(currow)] = 'Г'
                    else:
                        fnd2 = second_soup.find(text="Тип")
                        if fnd2:
                            ws['P' + str(currow)] = 'К'
                        else:
                            fnd3 = second_soup.find(text="Тип")
                            if fnd3:
                                ws['P' + str(currow)] = 'В'



                    # get type of o
                    fnd = second_soup.find(text="Тип")
                    if fnd:
                        print(fnd)
                        ws['I' + str(currow)] = 'Н'
                    else:
                        fnd2 = second_soup.find(text="Тип")
                        if fnd2:
                            ws['I' + str(currow)] = 'Г'

                    # parse parametrs of o
                    par_o = second_soup.find(text="Характеристики")
                    if par_o:
                        table_par_o = par_o.next.nextSibling
                        for par_o_str in table_par_o.find_all('tr'):
                            if par_o_str:
                                par_o_name = par_o_str.contents[1]
                                par_o_val = par_o_str.contents[2]
                                par_o_pos = find_charact_par_o(par_o_name.text)
                                if par_o_pos != '':
                                    ws[par_o_pos + str(currow)] = par_o_val.text
                                    print('    oit: ' + par_o_name.text + ' = ' + par_o_val.text)






                    #parse tech characteristics
                    tech = second_soup.find(text="Технические характеристики")
                    if tech:

                        table_tech = tech.next.nextSibling
                        for tech_str in table_tech.find_all('tr'):
                            if tech_str:
                                tech_name = tech_str.contents[1]
                                tech_val = tech_str.contents[2]
                                res = "     " + tech_name.text + " = " + tech_val.text
                                print(res)

                                if tech_name.text[:8] == 'М':
                                    ws['H' + str(currow-1)] = tech_val.text

                                tech_a = tech_str.find('a')
                                if tech_a:
                                    bool_items = True
                                    tech_href = tech_a.get('href')
                                    ws['T' + str(currow)] = tech_val.text  # name of item
                                    # go to the next inner level
                                    thirdpage = gethtml("http://www.......site.org/" + tech_href)
                                    third_soup = BeautifulSoup(thirdpage, 'html.parser')
                                    tech_in = third_soup.find(text="Технические характеристики")
                                    if tech_in:

                                        table_tech_in = tech_in.next.nextSibling
                                        for tech_str_in in table_tech_in.find_all('tr'):
                                            if  tech_str_in:
                                                tech_name_in = tech_str_in.contents[1]
                                                tech_val_in = tech_str_in.contents[2]
                                                res = "          " + tech_name_in.text + " = " + tech_val_in.text
                                                print(res)

                                                pos_char = find_charact(tech_name_in.text)
                                                if pos_char != '':
                                                    cur_txt = ''
                                                    upd_val = update_value_par(tech_name_in.text, tech_val_in.text)
                                                    if upd_val != tech_val_in.text:
                                                        print('************old = ' + tech_val_in.text + ' , new = ' + str(upd_val))
                                                    if ws[pos_char + str(currow)].value != None:
                                                        cur_txt = ws[pos_char + str(currow)].value+'\n'
                                                    ws[pos_char + str(currow)] = cur_txt + str(upd_val)
                                    currow += 1

                    if bool_items == True:
                        currow -= 1
                    bool_items = False

                    #Merge cells
                    Alph = string.ascii_uppercase[:-7]
                    for i in Alph:
                        a =i + str(rowbegin_1) + ':'+i + str(currow)
                        ws.merge_cells(i + str(rowbegin_1) + ':'+i + str(currow))

                    rowbegin_1 = currow+1

                    wb.save("Result.xlsx")

print("*******************************")
print("Parse is end")

