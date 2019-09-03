import psycopg2
from psycopg2 import sql
from openpyxl import load_workbook
import re



#------------------------------------connect to db------------------------------------------
def connectdb():
    try:
        conn = psycopg2.connect(dbname='db',
                                user='user',
                                password='pass',
                                host='localhost')
        print("Connect to database successfully")
        return conn
    except:
        print("Error while connecting to PostgreSQL")
        return None
#-------------------------------------------------------------------------------------------

#----------------------------Find value in column of table----------------------------------
def find_data_in_table(cursor, table, col, value, return_id):
    query_str = 'SELECT '+return_id+' FROM ' + table + ' WHERE '+col+' = %s;'
    cursor.execute(query_str, (value,))
    return(cursor.fetchone())
#-------------------------------------------------------------------------------------------


#-------------------------Insert values in columns of table---------------------------------
def insert_data_in_table(conn, cursor, table, cols, values, return_id):
    query_str = "INSERT INTO " + table + " (" + cols + ") VALUES ("+values+");  SELECT currval(pg_get_serial_sequence(%s, %s));"
    cursor.execute(query_str, (table, return_id,))
    res = cursor.fetchone()
    conn.commit()
    print('Add new value -'+ values+' into the table', table)
    return(res)
#-------------------------------------------------------------------------------------------


#---------------------------Find or insert and return id------------------------------------
def get_id_from_table(conn, cursor, table, col, value, id_return):
    find_id = find_data_in_table(cursor, table, col, value, id_return)
    if find_id is None:
        print(value, 'not found')
        cols = (col)
        values = ("'"+value+"'")
        find_id = insert_data_in_table(conn, cursor, table, cols, values, id_return)
    return find_id[0]
#-------------------------------------------------------------------------------------------


#-------------------------Update values in columns of table---------------------------------
def update_data_in_table_per_id(conn, cursor, table, col, value, return_id, value_id):
    query_str = "UPDATE " + table + " SET " + col + " = %s WHERE "+return_id+" = %s RETURNING "+return_id+";"
    cursor.execute(query_str, (value, value_id,))
    print('Update value '+ value + ' into the table', table)
    cursor.fetchone()
    conn.commit()
#-------------------------------------------------------------------------------------------



#--------------------------------Loading excel file-----------------------------------------
def load_excel(excel_name):
    try:
        wb = load_workbook(excel_name)
    except:
        print('Error with open file',excel_name)
        wb = None
    return wb
#-------------------------------------------------------------------------------------------


#------------------Preparing string from excel to save in postgresql------------------------
def prepare_str(txt):
    res = '-'
    if not (txt is None):
        txt = str(txt)
    if (txt != '') and (not (txt is None)):
        txt = txt.lstrip(' ')
        txt = re.sub("^\s+|\n|\r|\s+$", ' ', txt)
        txt = re.sub(" +", " ", txt)
        res = "".join(txt.split())
    if (res == '') or (txt is None) or (res is None):
        txt = '-'
    return txt
#-------------------------------------------------------------------------------------------


#------------------Preparing digit from excel to save in postgresql------------------------
def prepare_digit(txt):
    if txt.isdigit():
        return(int(txt))
    else:
        return(0)
#-------------------------------------------------------------------------------------------


#-------------------------Copy data from excel to postgresql--------------------------------
def excel_to_postgresql(conn, cursor, wb, ws):
    try:
        for i in range(4, ws.max_row+1):

            #Вид
            txt = prepare_str(ws['A'+str(i)].value)
            kind_id = get_id_from_table(conn, cursor,
                'slv_kind', 'name_kind', txt, 'id_kind')

            #Наименование (иностр.)----------------
            txt = prepare_str(ws['B'+str(i)].value)
            foreign_name_type_id = get_id_from_table(conn, cursor,
                'slv_name_type', 'name_type_foreign', txt, 'id_name_type')
            #Наименование (рос.)
            if (txt == '-'):
                txt = prepare_str(ws['C'+str(i)].value)
                rus_name_type_id = get_id_from_table(conn, cursor,
                    'slv_name_type_', 'name_type_rus', txt, 'id_name_type_')
            else:
                txt = prepare_str(ws['C'+str(i)].value)
                rus_name_type_id = foreign_name_type_id
                if (txt != '-'):
                    update_data_in_table_per_id(conn, cursor,
                        'slv_name_type_', 'name_type_rus', txt, 'id_name_type_', foreign_name_type_id)
            #-------------------------------------------------------

            #Примечание
            note = prepare_str(ws['D'+str(i)].value)

            #Класс 
            txt = prepare_str(ws['E'+str(i)].value)
            kind_id = get_id_from_table(conn, cursor,
                'slv_kind', 'name_kind', txt, 'id_kind')

            #Вид
            txt = prepare_str(ws['F'+str(i)].value)
            kind_a_id = get_id_from_table(conn, cursor,
                'slv_kind_a', 'name_kind_a', txt, 'id_kind_a')

            #Тип
            txt = prepare_str(ws['G'+str(i)].value)
            type_id = get_id_from_table(conn, cursor,
                'slv_type', 'name_type', txt, 'id_type')

            #Тип средств
            txt = prepare_str(ws['H'+str(i)].value)
            type_c_id = get_id_from_table(conn, cursor,
                'slv_type_c', 'name_type_c', txt, 'id_type_c')

            #Принадлежность
            owner_unit = prepare_str(ws['I'+str(i)].value)

            #Вербальное описание
            verb_using = prepare_str(ws['J'+str(i)].value)

            # i------------------------------------------
            txt = prepare_str(ws['K'+str(i)].value)
            ab_i_id = get_id_from_table(conn, cursor,
                'slv_i', 'code_i', txt, 'id_i')
            #п
            ab_name = prepare_str(ws['L'+str(i)].value)
            if (txt != '-') and (ab_name != '-'):
                update_data_in_table_per_id(conn, cursor,
                    'slv_i', 'name_i', txt, 'id_i', ab_i_id)
            #-------------------------------------------------------

            cols = ['kind_id',
                  'foreign_name_type_id',
                  'rus_name_type_id',
                  'note',
                  'kind_id',
                  'kind_a_id',
                  'type_id',
                  'type_c_id',
                  'owner_unit',
                  'verb_using',
                  'ab_i_id',
                  'ab_name']
            cols = ','.join(cols)

            values =[str(kind_id),
                  str(foreign_name_type_id),
                  str(rus_name_type_id),
                  "'" + note + "'",
                  str(kind_id),
                  str(kind_a_id),
                  str(type_id),
                  str(type_c_id),
                  "'" + owner_unit + "'",
                  "'" + verb_using + "'",
                  str(ab_i_id),
                  "'" + ab_name + "'"]
            values = ','.join(values)
            insert_data_in_table(conn, cursor, 'table_is', cols, values, 'id_is')
        print('Data successfully copied from excel to database')
    except:
        print('Error in copying from excel to database')
#-------------------------------------------------------------------------------------------






#------------------------------------------Main---------------------------------------------
def main(excel_name, sheet_name):
    conn = connectdb()
    if not (conn is None):
        cursor = conn.cursor()
        if cursor:
            print("Successfully create cursor")
            wb = load_excel(excel_name)
            if not (wb is None):
                ws = wb[sheet_name]
                print("Successfully open excel file")

                excel_to_postgresql(conn, cursor, wb, ws)

            else:
                print("Error with open excel file")
        else:
            print("Error with create cursor")
    else:
        print('Error with connection to database')

    if (cursor):
        cursor.close()
        conn.close()
        print("PostgreSQL connection is closed")
#-------------------------------------------------------------------------------------------






if __name__ == '__main__':
    main('work.xlsx', 'Infa')
